#!/usr/bin/env python3
"""
Generate BOM Excel workbook from Odoo 18 POS products.
Each product with a BOM gets its own sheet.
"""

import os
import sys
import json
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Odoo connection settings - adjust these
ODOO_URL = os.environ.get('ODOO_URL', 'http://localhost:8069')
ODOO_DB = os.environ.get('ODOO_DB', 'odoo')
ODOO_USER = os.environ.get('ODOO_USER', 'admin')
ODOO_PASSWORD = os.environ.get('ODOO_PASSWORD', 'admin')

class OdooClient:
    def __init__(self, url, db, user, password):
        self.url = url.rstrip('/')
        self.db = db
        self.user = user
        self.password = password
        self.uid = None
        self.session = requests.Session()

    def authenticate(self):
        """Authenticate with Odoo"""
        response = self.session.post(
            f"{self.url}/web/session/authenticate",
            json={
                "jsonrpc": "2.0",
                "method": "call",
                "params": {
                    "db": self.db,
                    "login": self.user,
                    "password": self.password
                },
                "id": 1
            }
        )
        result = response.json()
        if result.get('result', {}).get('uid'):
            self.uid = result['result']['uid']
            print(f"✓ Authenticated as {self.user} (uid: {self.uid})")
            return True
        else:
            print(f"✗ Authentication failed: {result.get('error', {}).get('message', 'Unknown error')}")
            return False

    def call(self, model, method, args=None, kwargs=None):
        """Call Odoo RPC method"""
        response = self.session.post(
            f"{self.url}/web/dataset/call_kw",
            json={
                "jsonrpc": "2.0",
                "method": "call",
                "params": {
                    "model": model,
                    "method": method,
                    "args": args or [],
                    "kwargs": kwargs or {}
                },
                "id": 2
            }
        )
        result = response.json()
        if 'error' in result:
            raise Exception(result['error'].get('message', str(result['error'])))
        return result.get('result')

    def search_read(self, model, domain=None, fields=None, limit=None, order=None):
        """Search and read records"""
        kwargs = {'domain': domain or []}
        if fields:
            kwargs['fields'] = fields
        if limit:
            kwargs['limit'] = limit
        if order:
            kwargs['order'] = order
        return self.call(model, 'search_read', kwargs=kwargs)


def get_pos_products_with_bom(client):
    """Get all POS products that have a BOM"""
    # Get products available in POS
    products = client.search_read(
        'product.product',
        domain=[('available_in_pos', '=', True), ('active', '=', True)],
        fields=['id', 'name', 'display_name', 'default_code', 'uom_id'],
        order='name'
    )
    print(f"Found {len(products)} POS products")

    # Get BOMs for these products
    product_ids = [p['id'] for p in products]
    boms = client.search_read(
        'mrp.bom',
        domain=[('product_id', 'in', product_ids), ('active', '=', True)],
        fields=['id', 'product_id', 'product_tmpl_id', 'product_qty', 'product_uom_id', 'bom_line_ids']
    )
    print(f"Found {len(boms)} BOMs")

    # Create a map of product_id -> BOM
    bom_map = {}
    for bom in boms:
        if bom['product_id']:
            bom_map[bom['product_id'][0]] = bom

    # Get BOM lines
    bom_line_ids = []
    for bom in boms:
        bom_line_ids.extend(bom.get('bom_line_ids', []))

    if bom_line_ids:
        bom_lines = client.search_read(
            'mrp.bom.line',
            domain=[('id', 'in', bom_line_ids)],
            fields=['id', 'bom_id', 'product_id', 'product_qty', 'product_uom_id']
        )

        # Group lines by BOM
        lines_by_bom = {}
        for line in bom_lines:
            bom_id = line['bom_id'][0]
            if bom_id not in lines_by_bom:
                lines_by_bom[bom_id] = []
            lines_by_bom[bom_id].append(line)
    else:
        lines_by_bom = {}

    # Build result
    result = []
    for product in products:
        if product['id'] in bom_map:
            bom = bom_map[product['id']]
            result.append({
                'product': product,
                'bom': bom,
                'lines': lines_by_bom.get(bom['id'], [])
            })

    print(f"Found {len(result)} products with BOMs")
    return result


def create_bom_sheet(ws, product_name, bom_lines):
    """Create a BOM sheet following the template format"""
    # Set column widths
    ws.column_dimensions['A'].width = 47
    ws.column_dimensions['B'].width = 35.5
    ws.column_dimensions['C'].width = 20

    # Styles
    title_font = Font(bold=True, size=26)
    header_font = Font(bold=True, size=24)
    data_font = Font(size=18)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: Title (merged A1:C1)
    ws.merge_cells('A1:C1')
    ws['A1'] = '物料清单（BOM）概述'
    ws['A1'].font = title_font
    ws['A1'].alignment = left_align
    ws.row_dimensions[1].height = 48

    # Row 2: Product name (merged A2:C2)
    ws.merge_cells('A2:C2')
    ws['A2'] = product_name
    ws['A2'].font = title_font
    ws['A2'].alignment = left_align
    ws.row_dimensions[2].height = 48

    # Row 3: Empty
    ws.row_dimensions[3].height = 20

    # Row 4: Headers
    headers = ['产品/物料', '数量', '单位']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 48

    # Row 5+: BOM lines
    for i, line in enumerate(bom_lines, 5):
        component_name = line.get('product_id', ['', '未知'])[1] if isinstance(line.get('product_id'), list) else str(line.get('product_id', ''))
        quantity = line.get('product_qty', 0)
        uom_name = line.get('product_uom_id', ['', ''])[1] if isinstance(line.get('product_uom_id'), list) else ''

        # Component name
        cell_a = ws.cell(row=i, column=1, value=component_name)
        cell_a.font = data_font
        cell_a.alignment = left_align
        cell_a.border = thin_border

        # Quantity
        cell_b = ws.cell(row=i, column=2, value=quantity)
        cell_b.font = data_font
        cell_b.alignment = center_align
        cell_b.border = thin_border

        # Unit
        cell_c = ws.cell(row=i, column=3, value=uom_name)
        cell_c.font = data_font
        cell_c.alignment = center_align
        cell_c.border = thin_border

        ws.row_dimensions[i].height = 48


def sanitize_sheet_name(name):
    """Sanitize sheet name for Excel (max 31 chars, no special chars)"""
    # Remove invalid characters
    invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        name = name.replace(char, '_')
    # Truncate to 31 characters
    if len(name) > 31:
        name = name[:28] + '...'
    return name


def main():
    """Main function"""
    print("=" * 60)
    print("BOM Excel Generator for Odoo 18 POS Products")
    print("=" * 60)

    # Check for custom Odoo settings from command line or env
    odoo_url = sys.argv[1] if len(sys.argv) > 1 else ODOO_URL
    odoo_db = sys.argv[2] if len(sys.argv) > 2 else ODOO_DB
    odoo_user = sys.argv[3] if len(sys.argv) > 3 else ODOO_USER
    odoo_password = sys.argv[4] if len(sys.argv) > 4 else ODOO_PASSWORD

    print(f"\nConnecting to: {odoo_url}")
    print(f"Database: {odoo_db}")
    print(f"User: {odoo_user}")

    # Connect to Odoo
    client = OdooClient(odoo_url, odoo_db, odoo_user, odoo_password)
    if not client.authenticate():
        sys.exit(1)

    # Get POS products with BOMs
    print("\nFetching POS products with BOMs...")
    products_with_bom = get_pos_products_with_bom(client)

    if not products_with_bom:
        print("No POS products with BOMs found!")
        sys.exit(0)

    # Create workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Create a sheet for each product
    for item in products_with_bom:
        product = item['product']
        bom_lines = item['lines']

        product_name = product.get('display_name') or product.get('name', 'Unknown')
        sheet_name = sanitize_sheet_name(product_name)

        print(f"  Creating sheet: {sheet_name} ({len(bom_lines)} components)")

        ws = wb.create_sheet(title=sheet_name)
        create_bom_sheet(ws, product_name, bom_lines)

    # Save workbook
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_file = f"/Users/taozhang/Downloads/POS_BOM_{timestamp}.xlsx"
    wb.save(output_file)

    print(f"\n✓ Generated: {output_file}")
    print(f"  Total sheets: {len(wb.sheetnames)}")
    print("\nDone!")


if __name__ == '__main__':
    main()
