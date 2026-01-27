# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class HrInsuranceRateImportWizard(models.TransientModel):
    _name = 'hr.insurance.rate.import.wizard'
    _description = 'Insurance Rate Import Wizard'

    file = fields.Binary(string='Excel File', required=True)
    filename = fields.Char(string='Filename')
    fiscal_year = fields.Char(string='Fiscal Year / 年度', required=True,
                               default=lambda self: self._default_fiscal_year())
    effective_date = fields.Date(string='Effective Date / 適用開始日', required=True)
    expire_previous = fields.Boolean(string='Expire Previous Rates / 旧料率を終了', default=True)
    preview_ids = fields.One2many('hr.insurance.rate.import.preview', 'wizard_id',
                                   string='Preview')

    @api.model
    def _default_fiscal_year(self):
        """Default to current Japanese fiscal year (Reiwa)"""
        from datetime import date
        today = date.today()
        # Japanese fiscal year starts April 1
        year = today.year if today.month >= 4 else today.year - 1
        # Reiwa era started May 1, 2019 (Reiwa 1)
        reiwa_year = year - 2018
        return f'R{reiwa_year}'

    def action_preview(self):
        """Parse Excel file and show preview"""
        self.ensure_one()

        if not OPENPYXL_AVAILABLE:
            raise UserError(_('openpyxl library is required. Install with: pip install openpyxl'))

        if not self.file:
            raise UserError(_('Please upload an Excel file.'))

        # Clear existing preview
        self.preview_ids.unlink()

        # Parse Excel file
        file_data = base64.b64decode(self.file)
        workbook = openpyxl.load_workbook(io.BytesIO(file_data), data_only=True)

        Prefecture = self.env['hr.insurance.prefecture']
        Preview = self.env['hr.insurance.rate.import.preview']

        preview_data = []
        for sheet_name in workbook.sheetnames:
            # Find prefecture by name
            prefecture = Prefecture.search([('name', '=', sheet_name)], limit=1)
            if not prefecture:
                continue

            sheet = workbook[sheet_name]

            # Parse health insurance rate (usually in cell C6 or similar)
            # This may need adjustment based on actual Excel structure
            health_rate = self._extract_rate(sheet, 'health')
            health_rate_care = self._extract_rate(sheet, 'health_care')

            preview_data.append({
                'wizard_id': self.id,
                'prefecture_id': prefecture.id,
                'health_rate': health_rate,
                'health_rate_with_care': health_rate_care,
                'pension_rate': 0.183,  # Fixed rate
                'employment_rate': 0.006,  # Fixed rate
            })

        if not preview_data:
            raise UserError(_('No valid prefecture data found in the Excel file.'))

        Preview.create(preview_data)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'hr.insurance.rate.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def _extract_rate(self, sheet, rate_type):
        """Extract rate from sheet based on rate type"""
        # Default implementation - may need customization based on Excel structure
        # Looking for patterns like "9.91%" or decimal values
        for row in range(1, 20):
            for col in range(1, 10):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    # Try to find percentage values
                    val = str(cell.value)
                    if '%' in val or (isinstance(cell.value, (int, float)) and 0 < cell.value < 1):
                        if isinstance(cell.value, (int, float)):
                            return cell.value if cell.value < 1 else cell.value / 100
        return 0.0

    def action_import(self):
        """Import rates from preview"""
        self.ensure_one()

        if not self.preview_ids:
            raise UserError(_('Please preview the data first.'))

        Rate = self.env['hr.insurance.rate']

        # Expire previous rates if requested
        if self.expire_previous:
            previous_rates = Rate.search([
                ('expiry_date', '=', False),
                ('effective_date', '<', self.effective_date),
            ])
            previous_rates.write({
                'expiry_date': self.effective_date,
            })

        # Create new rates
        created_rates = Rate
        for preview in self.preview_ids:
            rate = Rate.create({
                'prefecture_id': preview.prefecture_id.id,
                'fiscal_year': self.fiscal_year,
                'effective_date': self.effective_date,
                'health_rate': preview.health_rate,
                'health_rate_with_care': preview.health_rate_with_care,
                'pension_rate': preview.pension_rate,
                'employment_rate': preview.employment_rate,
            })
            created_rates |= rate

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Complete'),
                'message': _('%d insurance rates imported successfully.') % len(created_rates),
                'type': 'success',
                'sticky': False,
            }
        }


class HrInsuranceRateImportPreview(models.TransientModel):
    _name = 'hr.insurance.rate.import.preview'
    _description = 'Insurance Rate Import Preview'

    wizard_id = fields.Many2one('hr.insurance.rate.import.wizard', string='Wizard',
                                 required=True, ondelete='cascade')
    prefecture_id = fields.Many2one('hr.insurance.prefecture', string='Prefecture',
                                     required=True)
    health_rate = fields.Float(string='Health Rate', digits=(6, 5))
    health_rate_with_care = fields.Float(string='Health + Care Rate', digits=(6, 5))
    pension_rate = fields.Float(string='Pension Rate', digits=(6, 5))
    employment_rate = fields.Float(string='Employment Rate', digits=(6, 5))
