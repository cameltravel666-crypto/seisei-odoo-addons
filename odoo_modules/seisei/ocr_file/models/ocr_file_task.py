# -*- coding: utf-8 -*-

import json
import base64
import logging
import requests
import io
import os
import re
import time
from datetime import datetime

from odoo import models, fields, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# Central OCR Service Configuration (managed by Odoo 19)
OCR_SERVICE_URL = os.getenv('OCR_SERVICE_URL', 'https://ocr.seisei.tokyo/api/v1')
OCR_SERVICE_KEY = os.getenv('OCR_SERVICE_KEY', '')


def pdf_to_images(pdf_data: bytes, dpi: int = 150):
    """Convert PDF to images using PyMuPDF"""
    try:
        import fitz  # PyMuPDF
    except ImportError:
        raise ImportError('PyMuPDF required. Install: pip install PyMuPDF')

    images = []
    pdf_doc = fitz.open(stream=pdf_data, filetype='pdf')

    for page_num in range(len(pdf_doc)):
        page = pdf_doc[page_num]
        mat = fitz.Matrix(dpi / 72, dpi / 72)
        pix = page.get_pixmap(matrix=mat)
        img_data = pix.tobytes('jpeg')
        images.append(img_data)

    pdf_doc.close()
    return images


def build_ocr_prompt(template_fields: list) -> str:
    """Build OCR prompt based on template fields"""
    if not template_fields:
        # Default fields if no template headers found
        template_fields = [
            '仕入先名/Vendor Name',
            '日付/Date',
            '請求書番号/Invoice Number',
            '登録番号/Tax ID',
            '合計金額/Total Amount',
            '通貨/Currency'
        ]

    fields_json = {}
    for field in template_fields:
        # Clean field name for JSON key
        key = field.strip().replace(' ', '_').replace('/', '_')
        fields_json[key] = f"<{field}の値>"

    fields_str = json.dumps(fields_json, ensure_ascii=False, indent=2)

    prompt = f'''あなたは請求書・納品書のOCR専門家です。
この画像から以下の項目を抽出してJSON形式で返してください：

{fields_str}

また、明細行がある場合は以下の形式で抽出：
{{
  "line_items": [
    {{
      "product_name": "商品名",
      "quantity": 数量,
      "unit": "単位",
      "unit_price": 単価,
      "amount": 金額
    }}
  ]
}}

重要：
- 項目名は上記のキー名をそのまま使用
- 数値は数字のみ（通貨記号やカンマなし）
- 日付はYYYY-MM-DD形式
- 読み取れない項目はnullにする
- JSONのみを返す（説明文不要）'''

    return prompt


def process_image_via_central_service(image_data: bytes, mimetype: str = 'image/jpeg', template_fields: list = None, tenant_id: str = None, max_retries: int = 3):
    """Process image via Central OCR Service (managed by Odoo 19)

    This function does NOT directly call any AI API.
    All API calls are handled by the central service which:
    - Manages API keys securely
    - Tracks usage per tenant
    - Handles rate limiting and retries
    """
    b64_data = base64.standard_b64encode(image_data).decode('utf-8')

    if mimetype not in ['image/jpeg', 'image/png', 'image/gif', 'image/webp']:
        mimetype = 'image/jpeg'

    # Build request for central OCR service
    payload = {
        'image_data': b64_data,
        'mime_type': mimetype,
        'template_fields': template_fields or [],
        'tenant_id': tenant_id or 'default',
    }

    headers = {
        'Content-Type': 'application/json',
    }
    if OCR_SERVICE_KEY:
        headers['X-Service-Key'] = OCR_SERVICE_KEY

    last_error = None
    for attempt in range(max_retries):
        try:
            response = requests.post(
                f'{OCR_SERVICE_URL}/ocr/process',
                json=payload,
                headers=headers,
                timeout=120
            )

            if response.status_code == 503:
                # Service temporarily unavailable - wait and retry
                wait_time = (attempt + 1) * 5
                _logger.warning(f"[OCR] Service unavailable, waiting {wait_time}s before retry {attempt + 1}/{max_retries}")
                time.sleep(wait_time)
                last_error = 'service_unavailable'
                continue

            if response.status_code == 429:
                # Quota exceeded
                return {'success': False, 'error_code': 'quota_exceeded'}

            if response.status_code == 401:
                return {'success': False, 'error_code': 'auth_failed'}

            if response.status_code != 200:
                _logger.error(f"[OCR] Service error: {response.status_code}")
                return {'success': False, 'error_code': 'service_error'}

            result = response.json()

            if result.get('success'):
                return {
                    'success': True,
                    'extracted': result.get('extracted', {}),
                    'raw_response': result.get('raw_response', ''),
                    'usage': result.get('usage', {}),
                }
            else:
                return {'success': False, 'error_code': result.get('error_code', 'processing_failed')}

        except requests.exceptions.Timeout:
            last_error = 'timeout'
            if attempt < max_retries - 1:
                time.sleep(3)
                continue
            return {'success': False, 'error_code': 'timeout'}
        except requests.exceptions.ConnectionError:
            last_error = 'connection_error'
            if attempt < max_retries - 1:
                time.sleep(3)
                continue
            return {'success': False, 'error_code': 'connection_error'}
        except Exception as e:
            _logger.exception(f"[OCR] Unexpected error: {e}")
            return {'success': False, 'error_code': 'unexpected_error'}

    return {'success': False, 'error_code': last_error or 'max_retries'}


def get_user_friendly_error(error_code: str) -> str:
    """Convert error codes to user-friendly messages (no technical details)"""
    error_messages = {
        'quota_exceeded': 'OCR quota exceeded. Please contact administrator.',
        'auth_failed': 'OCR service authentication failed. Please contact administrator.',
        'service_error': 'OCR service is temporarily unavailable. Please try again later.',
        'service_unavailable': 'OCR service is temporarily unavailable. Please try again later.',
        'timeout': 'OCR processing timed out. Please try again with a smaller file.',
        'connection_error': 'Cannot connect to OCR service. Please try again later.',
        'processing_failed': 'Failed to process the document. Please ensure the image is clear.',
        'max_retries': 'OCR service is busy. Please try again later.',
        'unexpected_error': 'An error occurred. Please try again later.',
    }
    return error_messages.get(error_code, 'OCR processing failed. Please try again later.')


def _extract_json(text: str) -> dict:
    """Extract JSON from text response"""
    json_match = re.search(r'```(?:json)?\s*([\s\S]*?)```', text)
    if json_match:
        try:
            return json.loads(json_match.group(1))
        except json.JSONDecodeError:
            pass

    try:
        start = text.find('{')
        end = text.rfind('}') + 1
        if start != -1 and end > start:
            return json.loads(text[start:end])
    except json.JSONDecodeError:
        pass

    return {'raw_text': text}


def extract_template_headers(template_data: bytes) -> tuple:
    """
    Extract column headers from Excel template by auto-detecting header row.
    Returns (headers_list, header_row_number)
    """
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(template_data), read_only=True)
        sheet = workbook.active

        # Keywords that indicate a header row
        header_keywords = [
            '日付', '日期', 'date', '請求', '番号', 'number', 'no',
            '金額', '合計', 'total', 'amount', '仕入', 'vendor', 'supplier',
            '品名', '商品', 'product', 'item', '名前', 'name',
            '数量', 'quantity', 'qty', '単価', 'price', 'unit',
            '税', 'tax', '通貨', 'currency', 'ファイル', 'file',
            '備考', 'note', 'memo', '摘要', 'description'
        ]

        best_row = 1
        best_score = 0
        best_headers = []

        # Scan first 20 rows to find the header row
        for row_idx in range(1, min(21, sheet.max_row + 1)):
            row_cells = list(sheet[row_idx])

            # Count non-empty cells
            non_empty = [c for c in row_cells if c.value and str(c.value).strip()]
            if len(non_empty) < 2:
                continue

            # Calculate score based on header-like characteristics
            score = 0
            headers = []

            for cell in row_cells:
                if cell.value:
                    val = str(cell.value).strip().lower()
                    headers.append(str(cell.value).strip())

                    # Check if contains header keywords
                    for keyword in header_keywords:
                        if keyword.lower() in val:
                            score += 10
                            break

                    # Text cells are more likely headers than numbers
                    if not val.replace('.', '').replace(',', '').replace('-', '').isdigit():
                        score += 2

                    # Short text is more likely a header
                    if len(val) < 20:
                        score += 1
                else:
                    headers.append('')

            # Bonus for having multiple columns
            score += len(non_empty) * 2

            if score > best_score:
                best_score = score
                best_row = row_idx
                best_headers = [h for h in headers if h]  # Filter empty

        workbook.close()
        _logger.info(f"[OCR File] Detected header row {best_row}: {best_headers}")
        return best_headers, best_row

    except Exception as e:
        _logger.warning(f"[OCR File] Failed to extract headers: {e}")
        return [], 1


class OcrFileSource(models.Model):
    """Source file for OCR processing - supports multiple files"""
    _name = 'ocr.file.source'
    _description = 'OCR Source File'
    _order = 'sequence, id'

    task_id = fields.Many2one(
        'ocr.file.task',
        string='Task',
        required=True,
        ondelete='cascade'
    )
    sequence = fields.Integer(string='Sequence', default=10)

    # Source File
    source_file = fields.Binary(
        string='Source File',
        required=True,
        attachment=True,
        help='Image or PDF file to be processed by OCR'
    )
    source_filename = fields.Char(string='Filename')

    # OCR Status for this file
    state = fields.Selection([
        ('pending', 'Pending'),
        ('processing', 'Processing'),
        ('done', 'Done'),
        ('failed', 'Failed'),
    ], string='Status', default='pending')

    ocr_error_message = fields.Text(string='Error Message')

    # Extracted data - stored as JSON for flexibility
    ocr_raw_data = fields.Text(string='OCR Raw Data')
    extracted_data = fields.Text(string='Extracted Data (JSON)')

    # Common fields for quick access
    extracted_invoice_date = fields.Date(string='Invoice Date')
    extracted_invoice_number = fields.Char(string='Invoice Number')
    extracted_vendor_name = fields.Char(string='Vendor Name')
    extracted_tax_id = fields.Char(string='Tax ID')
    extracted_total_amount = fields.Float(string='Total Amount', digits=(16, 2))
    extracted_currency = fields.Char(string='Currency', default='JPY')
    extracted_line_items = fields.Text(string='Line Items (JSON)')


class OcrFileTask(models.Model):
    _name = 'ocr.file.task'
    _description = 'OCR File Processing Task'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'create_date desc'

    name = fields.Char(
        string='Task Name',
        required=True,
        default=lambda self: self._generate_task_name(),
        tracking=True
    )

    # === Template File ===
    template_file = fields.Binary(
        string='Excel Template',
        required=True,
        help='Upload the Excel template file (.xlsx) to be filled'
    )
    template_filename = fields.Char(string='Template Filename')
    template_headers = fields.Text(string='Template Headers (JSON)', readonly=True)
    template_header_row = fields.Integer(string='Header Row', default=1, readonly=True)

    # === Source Files (Multiple) ===
    source_file_ids = fields.One2many(
        'ocr.file.source',
        'task_id',
        string='Source Files',
        help='Upload invoice images or PDFs to be processed by OCR'
    )

    source_file_count = fields.Integer(
        string='File Count',
        compute='_compute_source_file_count',
        store=True
    )

    # === OCR Status ===
    state = fields.Selection([
        ('draft', 'Draft'),
        ('processing', 'Processing'),
        ('done', 'Done'),
        ('failed', 'Failed'),
    ], string='Status', default='draft', tracking=True)

    ocr_error_message = fields.Text(string='Error Message')
    ocr_processed_at = fields.Datetime(string='Processed At')

    # === Output File ===
    output_file = fields.Binary(string='Filled Template')
    output_filename = fields.Char(string='Output Filename')

    # === Computed Fields ===
    has_output = fields.Boolean(compute='_compute_has_output', store=False)

    # Usage info (computed)
    usage_this_month = fields.Integer(string='Used This Month', compute='_compute_usage_info')
    usage_remaining_free = fields.Integer(string='Free Remaining', compute='_compute_usage_info')
    usage_display = fields.Char(string='Usage', compute='_compute_usage_info')

    @api.depends('source_file_ids')
    def _compute_source_file_count(self):
        for record in self:
            record.source_file_count = len(record.source_file_ids)

    def _compute_usage_info(self):
        Usage = self.env['ocr.file.usage']
        for record in self:
            usage = Usage.get_current_usage()
            record.usage_this_month = usage.image_count
            record.usage_remaining_free = usage.remaining_free
            if usage.remaining_free > 0:
                record.usage_display = f"Free: {usage.remaining_free}/30 remaining"
            else:
                record.usage_display = f"Used: {usage.image_count} (¥{usage.billable_count * 20} charge)"

    @api.depends('output_file')
    def _compute_has_output(self):
        for record in self:
            record.has_output = bool(record.output_file)

    def _generate_task_name(self):
        """Generate a default task name with timestamp"""
        return f"OCR-{datetime.now().strftime('%Y%m%d-%H%M%S')}"

    @api.onchange('template_file')
    def _onchange_template_file(self):
        """Extract headers when template is uploaded"""
        if self.template_file:
            try:
                template_data = base64.b64decode(self.template_file)
                headers, header_row = extract_template_headers(template_data)
                self.template_headers = json.dumps(headers, ensure_ascii=False)
                self.template_header_row = header_row
            except Exception as e:
                _logger.warning(f"[OCR File] Error extracting headers: {e}")

    def _get_template_fields(self):
        """Get template fields from stored headers"""
        if self.template_headers:
            try:
                return json.loads(self.template_headers)
            except:
                pass

        # Try to extract from template file
        if self.template_file:
            template_data = base64.b64decode(self.template_file)
            headers, header_row = extract_template_headers(template_data)
            if headers:
                self.template_headers = json.dumps(headers, ensure_ascii=False)
                self.template_header_row = header_row
                return headers

        return []

    # === Actions ===
    def action_start_ocr(self):
        """Process all source files with Gemini OCR"""
        self.ensure_one()

        if not self.template_file:
            raise UserError('Please upload an Excel template first!')
        if not self.source_file_ids:
            raise UserError('Please upload at least one source document!')

        if self.state == 'processing':
            raise UserError('OCR is already processing.')

        # Validate template file type
        if self.template_filename and not self.template_filename.lower().endswith(('.xlsx', '.xls')):
            raise UserError('Template must be an Excel file (.xlsx or .xls)!')

        # Check quota
        pending_files = self.source_file_ids.filtered(lambda s: s.state != 'done')
        file_count = len(pending_files)

        Usage = self.env['ocr.file.usage']
        quota_info = Usage.check_quota(file_count)

        if quota_info['will_exceed'] and quota_info['new_billable'] > 0:
            # Show warning but continue processing
            _logger.info(f"[OCR File] Processing {file_count} files, {quota_info['new_billable']} billable at ¥{quota_info['price_per_image']}/image")

        # Get template fields for dynamic extraction
        template_fields = self._get_template_fields()
        _logger.info(f"[OCR File] Using template fields: {template_fields}")

        self.write({'state': 'processing', 'ocr_error_message': False})

        success_count = 0
        error_messages = []
        file_index = 0

        for source in self.source_file_ids:
            if source.state == 'done':
                success_count += 1
                continue

            # Add delay between API calls to avoid rate limiting (2 seconds between files)
            if file_index > 0:
                time.sleep(2)
            file_index += 1

            source.write({'state': 'processing'})

            try:
                # Decode file data
                file_data = base64.b64decode(source.source_file)
                filename = source.source_filename or ''

                # Determine mimetype
                if filename.lower().endswith('.pdf'):
                    mimetype = 'application/pdf'
                elif filename.lower().endswith('.png'):
                    mimetype = 'image/png'
                else:
                    mimetype = 'image/jpeg'

                _logger.info(f"[OCR File] Processing {filename} ({mimetype})")

                # Get tenant ID for usage tracking
                tenant_id = self.env.cr.dbname

                # Process PDF or image via central OCR service
                if mimetype == 'application/pdf':
                    try:
                        images = pdf_to_images(file_data)
                        if not images:
                            raise Exception('PDF has no pages')

                        # Process first page via central service
                        result = process_image_via_central_service(
                            images[0], 'image/jpeg', template_fields, tenant_id
                        )
                    except ImportError as e:
                        _logger.error(f"[OCR] PDF processing dependency missing: {e}")
                        result = {'success': False, 'error_code': 'processing_failed'}
                else:
                    result = process_image_via_central_service(
                        file_data, mimetype, template_fields, tenant_id
                    )

                if result.get('success'):
                    extracted = result.get('extracted', {})

                    # Store all extracted data as JSON
                    update_vals = {
                        'state': 'done',
                        'ocr_raw_data': json.dumps(result, ensure_ascii=False, indent=2),
                        'extracted_data': json.dumps(extracted, ensure_ascii=False, indent=2),
                        'ocr_error_message': False,
                    }

                    # Map common fields
                    update_vals['extracted_vendor_name'] = self._find_field_value(extracted,
                        ['vendor_name', '仕入先名', '仕入先', 'supplier', 'vendor', '取引先'])
                    update_vals['extracted_invoice_number'] = self._find_field_value(extracted,
                        ['invoice_number', '請求書番号', '伝票番号', 'invoice_no', 'number'])
                    update_vals['extracted_tax_id'] = self._find_field_value(extracted,
                        ['tax_id', '登録番号', 'registration_number', '適格請求書発行事業者登録番号'])
                    update_vals['extracted_currency'] = self._find_field_value(extracted,
                        ['currency', '通貨']) or 'JPY'

                    # Total amount
                    total = self._find_field_value(extracted,
                        ['total', 'total_amount', '合計', '合計金額', '請求金額', 'amount'])
                    if total:
                        try:
                            update_vals['extracted_total_amount'] = float(str(total).replace(',', ''))
                        except:
                            pass

                    # Parse date
                    date_str = self._find_field_value(extracted,
                        ['date', '日付', 'invoice_date', '請求日', '発行日'])
                    if date_str:
                        try:
                            date_str = re.sub(r'[年月]', '-', str(date_str))
                            date_str = re.sub(r'日', '', date_str)
                            update_vals['extracted_invoice_date'] = date_str
                        except Exception:
                            pass

                    # Store line items
                    line_items = extracted.get('line_items')
                    if line_items:
                        update_vals['extracted_line_items'] = json.dumps(line_items, ensure_ascii=False)

                    source.write(update_vals)
                    success_count += 1
                    _logger.info(f"[OCR File] {filename} processed successfully")
                else:
                    # Get user-friendly error message (hide technical details)
                    error_code = result.get('error_code', 'unknown')
                    user_error = get_user_friendly_error(error_code)
                    source.write({
                        'state': 'failed',
                        'ocr_error_message': user_error,
                    })
                    error_messages.append(f"{filename}: {user_error}")
                    # Log technical details for admin debugging only
                    _logger.warning(f"[OCR] {filename} failed with code: {error_code}")

            except Exception as e:
                # Log technical details for admin only
                _logger.exception(f"[OCR] Error processing {source.source_filename}: {e}")
                # Show user-friendly message (no technical details)
                user_error = get_user_friendly_error('unexpected_error')
                source.write({
                    'state': 'failed',
                    'ocr_error_message': user_error,
                })
                error_messages.append(f"{source.source_filename}: {user_error}")

        # Update usage count for successfully processed files
        if success_count > 0:
            Usage = self.env['ocr.file.usage']
            Usage.increment_usage(success_count)
            _logger.info(f"[OCR File] Incremented usage by {success_count}")

        # Update task state
        if success_count == len(self.source_file_ids):
            self.write({
                'state': 'done',
                'ocr_processed_at': fields.Datetime.now(),
            })
        elif success_count > 0:
            self.write({
                'state': 'done',
                'ocr_processed_at': fields.Datetime.now(),
                'ocr_error_message': f'{len(error_messages)} file(s) failed',
            })
        else:
            self.write({
                'state': 'failed',
                'ocr_error_message': '; '.join(error_messages[:3]),
            })

        # Return action to reload the form
        message = f'{success_count} file(s) processed successfully.'
        msg_type = 'success'
        if error_messages:
            message = f'{success_count}/{len(self.source_file_ids)} files processed. {len(error_messages)} failed.'
            msg_type = 'warning'

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'ocr.file.task',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
            'context': {
                'notification': {
                    'title': 'OCR Completed',
                    'message': message,
                    'type': msg_type,
                }
            }
        }

    def _find_field_value(self, data: dict, possible_keys: list):
        """Find value in extracted data by trying multiple possible keys"""
        for key in possible_keys:
            # Direct match
            if key in data:
                return data[key]
            # Case-insensitive match
            for data_key in data.keys():
                if key.lower() in data_key.lower() or data_key.lower() in key.lower():
                    return data[data_key]
        return None

    def _convert_to_excel_type(self, value):
        """Convert value to proper Excel type (number, date, or string)"""
        if value is None:
            return None

        # Already a number
        if isinstance(value, (int, float)):
            return value

        # Date object - return as-is for Excel
        if hasattr(value, 'year'):  # date or datetime
            return value

        # Convert string to proper type
        if isinstance(value, str):
            value = value.strip()

            # Empty string
            if not value:
                return None

            # Try to convert to number
            # Remove common formatting characters
            clean_value = value.replace(',', '').replace('￥', '').replace('¥', '').replace('円', '').replace(' ', '')

            # Check if it's a negative number in parentheses like (1234)
            if clean_value.startswith('(') and clean_value.endswith(')'):
                clean_value = '-' + clean_value[1:-1]

            try:
                # Try integer first
                if '.' not in clean_value:
                    return int(clean_value)
                else:
                    return float(clean_value)
            except (ValueError, TypeError):
                pass

            # Try to parse as date (YYYY-MM-DD or similar)
            try:
                from datetime import datetime as dt
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日']:
                    try:
                        return dt.strptime(value, fmt).date()
                    except ValueError:
                        continue
            except:
                pass

            # Return as string
            return value

        return value

    def action_reset(self):
        """Reset task to draft state"""
        for source in self.source_file_ids:
            source.write({
                'state': 'pending',
                'ocr_raw_data': False,
                'extracted_data': False,
                'ocr_error_message': False,
                'extracted_invoice_date': False,
                'extracted_invoice_number': False,
                'extracted_vendor_name': False,
                'extracted_tax_id': False,
                'extracted_total_amount': 0,
                'extracted_line_items': False,
            })

        self.write({
            'state': 'draft',
            'ocr_error_message': False,
            'ocr_processed_at': False,
            'output_file': False,
            'output_filename': False,
        })

    def action_fill_template(self):
        """Fill the Excel template with extracted OCR data from all files"""
        self.ensure_one()

        if self.state != 'done':
            raise UserError('OCR processing must be completed before filling the template!')

        if not self.template_file:
            raise UserError('Template file is missing!')

        done_sources = self.source_file_ids.filtered(lambda s: s.state == 'done')
        if not done_sources:
            raise UserError('No successfully processed files to fill template!')

        try:
            from openpyxl import load_workbook

            # Load template
            template_data = base64.b64decode(self.template_file)
            workbook = load_workbook(io.BytesIO(template_data))
            sheet = workbook.active

            # Get header row (auto-detected or default to 1)
            header_row = self.template_header_row or 1

            # Get headers from detected header row
            headers = []
            for cell in sheet[header_row]:
                headers.append(str(cell.value).strip() if cell.value else '')

            # Data starts from the row after header
            start_row = header_row + 1

            # Check for {{data_start}} marker override
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20), start=1):
                for cell in row:
                    if cell.value == '{{data_start}}':
                        start_row = row_idx + 1
                        cell.value = None
                        break

            # Get merged cell ranges to avoid writing to merged cells
            from openpyxl.cell.cell import MergedCell

            def is_merged_cell(sheet, row, col):
                """Check if a cell is part of a merged range (but not the top-left)"""
                cell = sheet.cell(row=row, column=col)
                return isinstance(cell, MergedCell)

            def get_writable_cell(sheet, row, col):
                """Get the writable cell - for merged cells, find the top-left"""
                for merged_range in sheet.merged_cells.ranges:
                    if (merged_range.min_row <= row <= merged_range.max_row and
                        merged_range.min_col <= col <= merged_range.max_col):
                        # Return the top-left cell of the merged range
                        return sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                return sheet.cell(row=row, column=col)

            # Fill data from each source file
            current_row = start_row
            for source in done_sources:
                # Load extracted data
                extracted = {}
                if source.extracted_data:
                    try:
                        extracted = json.loads(source.extracted_data)
                    except:
                        pass

                # Fill each column based on header
                for col_idx, header in enumerate(headers, start=1):
                    if not header:
                        continue

                    # Skip if this is a merged cell (not the primary cell)
                    if is_merged_cell(sheet, current_row, col_idx):
                        continue

                    # Find matching value in extracted data
                    value = self._find_field_value(extracted, [header])

                    # If not found, try common field mappings
                    if value is None:
                        header_lower = header.lower()
                        if 'vendor' in header_lower or '仕入先' in header:
                            value = source.extracted_vendor_name
                        elif 'date' in header_lower or '日付' in header:
                            value = source.extracted_invoice_date  # Keep as date object
                        elif 'invoice' in header_lower or '請求書' in header or '番号' in header:
                            value = source.extracted_invoice_number
                        elif 'tax' in header_lower or '登録' in header:
                            value = source.extracted_tax_id
                        elif 'total' in header_lower or '合計' in header or '金額' in header:
                            value = source.extracted_total_amount  # Already float
                        elif 'currency' in header_lower or '通貨' in header:
                            value = source.extracted_currency
                        elif 'file' in header_lower or 'ファイル' in header:
                            value = source.source_filename

                    if value is not None:
                        try:
                            # Convert to proper type for Excel
                            cell_value = self._convert_to_excel_type(value)
                            sheet.cell(row=current_row, column=col_idx, value=cell_value)
                        except AttributeError:
                            # Skip merged cells that can't be written
                            _logger.warning(f"[OCR File] Skipping merged cell at row={current_row}, col={col_idx}")
                            pass

                current_row += 1

            # Save output
            output_buffer = io.BytesIO()
            workbook.save(output_buffer)
            output_data = base64.b64encode(output_buffer.getvalue())

            # Generate output filename
            base_name = self.template_filename or 'output'
            if base_name.lower().endswith(('.xlsx', '.xls')):
                base_name = base_name[:-5] if base_name.lower().endswith('.xlsx') else base_name[:-4]
            output_filename = f"{base_name}_filled_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            self.write({
                'output_file': output_data,
                'output_filename': output_filename,
            })

            # Return action to reload form and show download section
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'ocr.file.task',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'current',
            }

        except ImportError:
            raise UserError('openpyxl library is required. Please install it: pip install openpyxl')
        except Exception as e:
            _logger.exception(f"[OCR File] Error filling template: {e}")
            raise UserError(f'Error filling template: {str(e)}')

    def action_download_output(self):
        """Download the filled template file"""
        self.ensure_one()

        if not self.output_file:
            raise UserError('No output file available. Please fill the template first.')

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content?model=ocr.file.task&id={self.id}&field=output_file&filename_field=output_filename&download=true',
            'target': 'self',
        }

    def action_add_files(self):
        """Open wizard to add multiple files"""
        return {
            'name': 'Add Source Files',
            'type': 'ir.actions.act_window',
            'res_model': 'ocr.file.upload.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_task_id': self.id},
        }
