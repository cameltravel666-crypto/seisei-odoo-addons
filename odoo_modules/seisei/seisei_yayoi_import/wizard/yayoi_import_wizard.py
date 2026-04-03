import base64
import csv
import io
import logging
import math

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Full-width katakana → half-width katakana mapping (copied from account_move)
# ---------------------------------------------------------------------------
_FULL_TO_HALF_KANA = {
    'ア': 'ｱ', 'イ': 'ｲ', 'ウ': 'ｳ', 'エ': 'ｴ', 'オ': 'ｵ',
    'カ': 'ｶ', 'キ': 'ｷ', 'ク': 'ｸ', 'ケ': 'ｹ', 'コ': 'ｺ',
    'サ': 'ｻ', 'シ': 'ｼ', 'ス': 'ｽ', 'セ': 'ｾ', 'ソ': 'ｿ',
    'タ': 'ﾀ', 'チ': 'ﾁ', 'ツ': 'ﾂ', 'テ': 'ﾃ', 'ト': 'ﾄ',
    'ナ': 'ﾅ', 'ニ': 'ﾆ', 'ヌ': 'ﾇ', 'ネ': 'ﾈ', 'ノ': 'ﾉ',
    'ハ': 'ﾊ', 'ヒ': 'ﾋ', 'フ': 'ﾌ', 'ヘ': 'ﾍ', 'ホ': 'ﾎ',
    'マ': 'ﾏ', 'ミ': 'ﾐ', 'ム': 'ﾑ', 'メ': 'ﾒ', 'モ': 'ﾓ',
    'ヤ': 'ﾔ', 'ユ': 'ﾕ', 'ヨ': 'ﾖ',
    'ラ': 'ﾗ', 'リ': 'ﾘ', 'ル': 'ﾙ', 'レ': 'ﾚ', 'ロ': 'ﾛ',
    'ワ': 'ﾜ', 'ヲ': 'ｦ', 'ン': 'ﾝ',
    'ァ': 'ｧ', 'ィ': 'ｨ', 'ゥ': 'ｩ', 'ェ': 'ｪ', 'ォ': 'ｫ',
    'ッ': 'ｯ', 'ャ': 'ｬ', 'ュ': 'ｭ', 'ョ': 'ｮ',
    'ー': 'ｰ', '゛': 'ﾞ', '゜': 'ﾟ',
    'ガ': 'ｶﾞ', 'ギ': 'ｷﾞ', 'グ': 'ｸﾞ', 'ゲ': 'ｹﾞ', 'ゴ': 'ｺﾞ',
    'ザ': 'ｻﾞ', 'ジ': 'ｼﾞ', 'ズ': 'ｽﾞ', 'ゼ': 'ｾﾞ', 'ゾ': 'ｿﾞ',
    'ダ': 'ﾀﾞ', 'ヂ': 'ﾁﾞ', 'ヅ': 'ﾂﾞ', 'デ': 'ﾃﾞ', 'ド': 'ﾄﾞ',
    'バ': 'ﾊﾞ', 'ビ': 'ﾋﾞ', 'ブ': 'ﾌﾞ', 'ベ': 'ﾍﾞ', 'ボ': 'ﾎﾞ',
    'パ': 'ﾊﾟ', 'ピ': 'ﾋﾟ', 'プ': 'ﾌﾟ', 'ペ': 'ﾍﾟ', 'ポ': 'ﾎﾟ',
    'ヴ': 'ｳﾞ',
}


def to_halfwidth_kana(text):
    """Convert full-width katakana to half-width for Yayoi compatibility."""
    if not text:
        return ''
    return ''.join(_FULL_TO_HALF_KANA.get(ch, ch) for ch in text)


def truncate_yayoi(text, max_halfwidth):
    """Truncate text to fit within *max_halfwidth* half-width character count.

    CJK / full-width characters count as 2 half-width units.
    """
    if not text:
        return ''
    text = text.replace('\n', '').replace('\r', '')
    width = 0
    result = []
    for ch in text:
        cp = ord(ch)
        cw = 2 if (
            (0x3000 <= cp <= 0x9FFF)
            or (0xF900 <= cp <= 0xFAFF)
            or (0xFF01 <= cp <= 0xFF60)
            or (0xFFE0 <= cp <= 0xFFE6)
            or (0x20000 <= cp <= 0x2FA1F)
        ) else 1
        if width + cw > max_halfwidth:
            break
        width += cw
        result.append(ch)
    return ''.join(result)


# ---------------------------------------------------------------------------
# TransientModel: wizard line
# ---------------------------------------------------------------------------

class YayoiImportLine(models.TransientModel):
    _name = 'seisei.yayoi.import.line'
    _description = 'Yayoi Import Line'
    _order = 'seq_no, id'

    wizard_id = fields.Many2one('seisei.yayoi.import.wizard', ondelete='cascade')
    date = fields.Date(string='日付')
    description = fields.Char(string='適用')
    account_name = fields.Char(string='勘定科目')
    receipt_no = fields.Char(string='領収書No.')
    seq_no = fields.Integer(string='伝票番号')
    amount = fields.Integer(string='金額')
    tax_rate = fields.Selection([
        ('0', '非課税'),
        ('8', '8%'),
        ('10', '10%'),
    ], string='税率')
    tax_amount = fields.Integer(string='税額', compute='_compute_tax', store=True)
    tax_category = fields.Char(string='税区分', compute='_compute_tax', store=True)

    @api.depends('amount', 'tax_rate')
    def _compute_tax(self):
        for line in self:
            rate = int(line.tax_rate or '0')
            if rate == 10:
                line.tax_amount = line.amount - int(line.amount * 100 / 110)
                line.tax_category = '課税売上込10%'
            elif rate == 8:
                line.tax_amount = line.amount - int(line.amount * 100 / 108)
                line.tax_category = '課税売上込軽減8%'
            else:
                line.tax_amount = 0
                line.tax_category = '対象外'


# ---------------------------------------------------------------------------
# TransientModel: main wizard
# ---------------------------------------------------------------------------

class YayoiImportWizard(models.TransientModel):
    _name = 'seisei.yayoi.import.wizard'
    _description = 'Yayoi Import Wizard'

    file_data = fields.Binary(string='Excel ファイル', required=True)
    filename = fields.Char(string='ファイル名')
    state = fields.Selection([
        ('upload', 'Upload'),
        ('preview', 'Preview'),
    ], default='upload', string='状態')
    line_ids = fields.One2many(
        'seisei.yayoi.import.line', 'wizard_id', string='明細',
    )
    debit_account = fields.Char(string='借方勘定科目', default='仕入高')
    credit_account = fields.Char(string='貸方勘定科目', default='現金')

    # Download helpers
    export_data = fields.Binary(string='Export Data', readonly=True)
    export_filename = fields.Char(string='Export Filename', readonly=True)

    # ---- action: parse Excel → preview lines ----

    def action_parse(self):
        self.ensure_one()
        if not self.file_data:
            raise UserError(_('ファイルをアップロードしてください。'))

        try:
            import openpyxl
        except ImportError:
            raise UserError(_('openpyxl がインストールされていません。'))

        raw = base64.b64decode(self.file_data)
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)

        # Try to find the sheet
        sheet_name = None
        for name in ('領収書2', '領収書'):
            if name in wb.sheetnames:
                sheet_name = name
                break
        if not sheet_name:
            raise UserError(
                _('シート「領収書2」または「領収書」が見つかりません。\nシート名: %s')
                % ', '.join(wb.sheetnames)
            )

        ws = wb[sheet_name]
        lines_vals = []
        for row_idx in range(4, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value   # A: 日付
            desc_val = ws.cell(row=row_idx, column=2).value    # B: 適用
            acct_val = ws.cell(row=row_idx, column=3).value    # C: Account (勘定科目)
            non_tax = ws.cell(row=row_idx, column=4).value     # D: 非課税
            rate_8 = ws.cell(row=row_idx, column=5).value      # E: 8%
            rate_10 = ws.cell(row=row_idx, column=6).value     # F: 10%
            subtotal = ws.cell(row=row_idx, column=7).value    # G: 小計
            receipt_no = ws.cell(row=row_idx, column=9).value  # I: 領収書No.
            seq_raw = ws.cell(row=row_idx, column=11).value    # K: 伝票番号

            # Skip rows where subtotal is 0 or missing
            if not subtotal or subtotal == 0:
                continue

            amount = int(subtotal)

            # Determine tax rate: only one of D/E/F is filled per row
            if rate_10 and float(rate_10) != 0:
                tax_rate = '10'
            elif rate_8 and float(rate_8) != 0:
                tax_rate = '8'
            else:
                tax_rate = '0'

            # Parse date
            parsed_date = False
            if date_val:
                if hasattr(date_val, 'date'):
                    parsed_date = date_val.date() if hasattr(date_val, 'date') else date_val
                elif hasattr(date_val, 'strftime'):
                    parsed_date = date_val
                else:
                    # Try string parse
                    try:
                        from datetime import datetime
                        parsed_date = datetime.strptime(str(date_val)[:10], '%Y-%m-%d').date()
                    except Exception:
                        parsed_date = False

            # Parse seq_no
            try:
                seq_no = int(seq_raw) if seq_raw else 0
            except (ValueError, TypeError):
                seq_no = 0

            lines_vals.append({
                'date': parsed_date,
                'description': str(desc_val or ''),
                'account_name': str(acct_val or ''),
                'receipt_no': str(receipt_no or ''),
                'seq_no': seq_no,
                'amount': amount,
                'tax_rate': tax_rate,
            })

        if not lines_vals:
            raise UserError(_('有効なデータ行がありませんでした。'))

        # Clear existing lines and create new ones
        self.line_ids.unlink()
        for vals in lines_vals:
            vals['wizard_id'] = self.id
        self.env['seisei.yayoi.import.line'].create(lines_vals)

        self.write({'state': 'preview'})

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    # ---- action: export Yayoi TXT ----

    def action_export_txt(self):
        self.ensure_one()
        if not self.line_ids:
            raise UserError(_('エクスポートする明細がありません。'))

        t = truncate_yayoi
        h = to_halfwidth_kana

        default_debit_acct = self.debit_account or '仕入高'
        credit_acct = self.credit_account or '現金'

        output = io.BytesIO()
        wrapper = io.TextIOWrapper(output, encoding='cp932', newline='')
        writer = csv.writer(wrapper, quoting=csv.QUOTE_NONNUMERIC)

        for line in self.line_ids.sorted(lambda l: (l.date or fields.Date.today(), l.seq_no)):
            date_str = ''
            if line.date:
                date_str = f'{line.date.year}/{line.date.month}/{line.date.day}'

            # Build memo: No.XXXX--MM/DD-description (same for col17 & col22)
            memo = line.receipt_no or ''

            rate = int(line.tax_rate or '0')

            # Debit account: use line-level account_name, fallback to wizard default
            debit_acct = line.account_name or default_debit_acct

            # Debit side: expense account with tax
            if rate == 10:
                debit_tax_cat = '課対仕入込10%'
            elif rate == 8:
                debit_tax_cat = '課対仕入込軽減8%'
            else:
                debit_tax_cat = '対象外'

            # Credit side: always 対象外 for cash
            credit_tax_cat = '対象外'

            row = [
                '2000',                         # Col1: 識別フラグ (str → quoted)
                line.seq_no or 0,               # Col2: 伝票番号 (int → unquoted)
                '',                             # Col3: 決算
                date_str,                       # Col4: 取引日付
                t(debit_acct, 24),              # Col5: 借方勘定科目
                '',                             # Col6: 借方補助科目
                '',                             # Col7: 借方部門
                debit_tax_cat,                  # Col8: 借方税区分
                line.amount,                    # Col9: 借方金額
                line.tax_amount,                # Col10: 借方税金額
                t(credit_acct, 24),             # Col11: 貸方勘定科目
                '',                             # Col12: 貸方補助科目
                '',                             # Col13: 貸方部門
                credit_tax_cat,                 # Col14: 貸方税区分
                line.amount,                    # Col15: 貸方金額
                0,                              # Col16: 貸方税金額
                t(h(memo), 64),                 # Col17: 摘要
                '',                             # Col18: 番号
                '',                             # Col19: 期日
                0,                              # Col20: タイプ (0=仕訳)
                '',                             # Col21: 生成元
                t(h(memo), 64),                 # Col22: 仕訳メモ
                '0',                            # Col23: 付箋1 (str → quoted)
                '0',                            # Col24: 付箋2 (str → quoted)
                'no',                           # Col25: 調整
            ]
            writer.writerow(row)

        wrapper.flush()
        wrapper.detach()

        filename = f'yayoi_import_{fields.Date.today().strftime("%Y%m%d")}.txt'
        self.write({
            'export_data': base64.b64encode(output.getvalue()),
            'export_filename': filename,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': (
                f'/web/content?model={self._name}'
                f'&id={self.id}'
                f'&field=export_data'
                f'&filename_field=export_filename'
                f'&download=true'
            ),
            'target': 'self',
        }

    # ---- action: go back to upload ----

    def action_back(self):
        self.ensure_one()
        self.write({'state': 'upload'})
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }
